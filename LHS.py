from PIL import Image as im
import string
import openpyxl
from openpyxl import Workbook
import requests
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Side
def coor(b,c):
    import string
    l=string.ascii_uppercase
    c=str(c)
    if b>25:
      b=l[-1+b//26]+l[b%26]
    else:
        b=l[b]
    x=b+c
    return x
def cif(b):
    if b>25:
      b=l[-1+b//26]+l[b%26]
    else:
        b=l[b]
    return b
borderf=Border(left=Side(border_style='thick',
                             color='00000000'),
                   right=Side(border_style='thin',
                              color='b0b0b0b0'),
                   top=Side(border_style='thick',
                             color='00000000'),
                   bottom=Side(border_style='thin',
                               color='b0b0b0b0'))
bordercl = Border(left=Side(border_style='thick',
                             color='00000000'))
bordercl2= Border(left=Side(border_style='thick',
                             color='00000000'),top=Side(border_style='thin',
                            color='b0b0b0b0'))
bordercv = Border(top=Side(border_style='thick',
                             color='00000000'))
bordercv2 = Border(top=Side(border_style='thick',
                             color='00000000'),left=Side(border_style='thin',
                             color='b0b0b0b0'))
border2 = Border(left=Side(border_style='thin',
                             color='b0b0b0b0'),
                   right=Side(border_style='thin',
                              color='b0b0b0b0'),
                   top=Side(border_style='thin',
                            color='b0b0b0b0'),
                   bottom=Side(border_style='thin',
                               color='b0b0b0b0'))
fill = PatternFill(fill_type='solid',
                   start_color='000000',
                   end_color='000000')
fillkv=PatternFill(fill_type='solid',
                   start_color='b0b0b0',
                   end_color='b0b0b0')
name=input('название будущего excel файла ')
name=name+'.xlsx'
Im=im.new('RGB',(1000,1000))
Im.save(name+'.png')
fak=input('URL картинки ')
lvr=int(input('длина рамки слева в клетках '))
lvo=int(input('отступ от начала картинки слева '))
lv=lvr+lvo
ver=int(input('длина рамки в клетках сверху '))
veo=int(input('отступ от начала картинки сверху '))
ve=ver+veo
gor=int(input('длина картинки в клетках по горизонстали '))
vert=int(input('длина картинки в клетках по вертикали '))
p = requests.get(fak)
out = open(name+'.png', "wb")
out.write(p.content)
out.close()
image = im.open(name+'.png')
width = image.size[0]
height = image.size[1]
print(width)
print(height)
pix=image.load()
rl=2
rv=2
rp=2
rn=2
pobeda=-1
for i in range(width):
    rl+=1
    if abs(pix[i+1,round(height/2)][0]-176)<10 and abs(pix[i,round(height/2)][0])<10:
        break
for i in range(height):
    rv+=1
    if abs(pix[round(width/2),i+1][0]-176)<10 and abs(pix[round(width/2),i][0])<10:
        break
for i in range(width):
    rp+=1
    if abs(pix[width-2-i,round(height/2)][0]-176)<10 and abs(pix[width-1-i,round(height/2)][0])<10:
        break
for i in range(height):
    rn+=1
    if abs(pix[round(width/2),height-2-i][0]-176)<10 and abs(pix[round(width/2),height-1-i][0])<10:
        break
area=(rl,rv,width-rp,height-rn)
image=image.crop(area)
#image.save(name+'1.png')
#image = im.open(name+'1.png')
image.show()
width = image.size[0]
height = image.size[1]
pix=image.load()
n=0
s=[]
sto=[]
nol=[]
dve=[]
c1=[0]
l=string.ascii_uppercase
for i in range(width):
 for j in range(height):
    if pix[i,j][0] not in s:
        s.append(pix[i,j][0])
        if abs(176-pix[i,j][0])<10:
            sto.append(pix[i,j][0])
        if abs(pix[i,j][0])<10:
            nol.append(pix[i,j][0])
        if abs(255-pix[i,j][0])<10:
            dve.append(pix[i,j][0])
for i in range(len(s)):
    if s[i] in sto or s[i] in nol or s[i] in dve:
        print('ok')
    else:
        print('sosi')
        print(s[i])
wb = Workbook()
sheet=wb.active
for i in range(width-1):
  c=[]
  c1=[0]
  if (pix[i+1,1][0] in sto and pix[i,j][0] not in sto) or (i+1==width-1 and pix[i,j][0] not in sto):
     pobeda+=1
     for j in range(height-1):
        if (pix[i,j+1][0] in sto and pix[i,j][0] in nol) or (j+1==height-1 and pix[i,j][0] in nol):
            c.append('ch')
        if (pix[i,j+1][0] in sto and pix[i,j][0] in dve) or (j+1==height-1 and pix[i,j][0] in dve):
            c.append('b')
     for f in range(len(c)):
         if c[f]=='ch':
             sheet[coor(pobeda+lv,f+ve+1)].fill=fill
             c1[len(c1)-1]+=1
         else:
             if c1!=[0] and c[f-1]=='ch':
                 c1.append(0)
     if c1[len(c1)-1]==0:
        c1.remove(0)
     #print(c1)
     #print(1)
     for i in range(len(c1)):
      sheet[coor(lv+n,ve-len(c1)+1+i)]=c1[i]
     n+=1
n=0
for i in range(height-1):
  j=0
  c=[]
  c1=[0]
  if (pix[1,i+1][0] in sto and pix[j,i][0] not in sto) or (i+1==height-1 and pix[j,i][0] not in sto):
     for j in range(width-1):
        if (pix[j+1,i][0] in sto and pix[j,i][0] in nol) or (j+1==width-1 and pix[j,i][0] in nol):
            c.append('ch')
        if (pix[j+1,i][0] in sto and pix[j,i][0] in dve) or (j+1==width-1 and pix[j,i][0] in dve):
            c.append('b')
     for f in range(len(c)):
         if c[f]=='ch':
             c1[len(c1)-1]+=1
         else:
             if c1!=[0] and c[f-1]=='ch':
                 c1.append(0)
     if c1[len(c1)-1]==0:
        c1.remove(0)
     #print(c1)
     for i in range(len(c1)):
      cif=str(ve+n+1)
      if lv-len(c1)>25:
       buk='A'+l[lv-len(c1)-26+i]
      else:
       buk=l[lv-len(c1)-26+i]
      sheet[buk+cif]=c1[i]
     n+=1
for cellObj in sheet[coor(lvo,veo+1)+':'+coor(lv+gor-1,ve+vert)]:
    for cell in cellObj:
        sheet[cell.coordinate].border = border2
for i in range(lvo,lv+gor):
    sheet[coor(i,veo+1)].border=bordercv2
    sheet[coor(i,veo+1+ver)].border=bordercv2
    sheet[coor(i,ve+vert+1)].border=bordercv
for i in range(veo,ve+vert):
    sheet[coor(lvo,i+1)].border=bordercl2
    sheet[coor(lvo+lvr,i+1)].border=bordercl2
    sheet[coor(lv+gor,i+1)].border=bordercl
sheet[coor(lvo,veo+1)].border=borderf
sheet[coor(lvo,ver+veo+1)].border=borderf
sheet[coor(lvr+lvo,veo+1)].border=borderf
sheet[coor(lvr+lvo,ver+veo+1)].border=borderf
for i in sheet[coor(lvo,veo+1)+':'+coor(lvr+lvo-1,ver+veo)]:
    for j in i:
        sheet[j.coordinate].fill=fillkv
for i in range(gor+lv):
    sheet.column_dimensions[sheet[1][i].column].width =3
wb.save(name)
