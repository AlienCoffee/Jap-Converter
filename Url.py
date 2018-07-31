# -*- coding: utf-8 -*-
#1.Preporation
#import
import os
from PIL import Image as im
import openpyxl
from openpyxl import Workbook
import requests
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Side
import time
# buk-выдаёт число(стобец) покоординате
#coor-выдаёт чисто в формате буквацифра('A1' )при вводе координат (0,1)
#border-рамка
#fill-заливка
#align_center-выравнивание
def buk(b):
    import string
    l=string.ascii_uppercase
    if b>25:
      b=l[-1+b//26]+l[b%26]
    else:
        b=l[b]
    return b
def coor(b,c):
    c=str(c)
    b=buk(b)
    c=b+c
    return c
def border(q,r):
    for i in range(4):
        if q[i]==1:
            q[i]='medium'
        elif q[i]==0:
            q[i]='thin'
        else:
            q[i]='medium'
        if r[i]==0:
            r[i]='00000000'
        elif r[i]==1:
            r[i]='00000000'
    border=Border(left=Side(border_style=q[0],
                             color=r[0]),
                   right=Side(border_style=q[1],
                              color=r[1]),
                   top=Side(border_style=q[2],
                             color=r[2]),
                   bottom=Side(border_style=q[3],
                               color=r[3]))
    return border
def fill(a):
    if a==0:
        a='444444'
    else:
        a='b0b0b0'
    fill=PatternFill(fill_type='solid',
                   start_color=a,
                   end_color=a)
    return fill
align_center=Alignment(horizontal='center',
                       vertical='center',
                       text_rotation=0,
                       wrap_text=False,
                       shrink_to_fit=False,
                       indent=0)
font = Font(name='Cambria',
                    size=15,
                    bold=False,
                    italic=False,
                    vertAlign=None,
                    underline='none',
                    strike=False,
                    color='00000000')
#2.Приём
name=input('название будущего excel файла ')
lvr=int(input('длина рамки (в клетках) слева '))
lvo=1
ver=int(input('длина рамки (в клетках) сверху '))
veo=4
fak=input('URL картинки ')
date=input('хотите использовать сегодняшнуюю дату?(1-да 2-нет) ')
if date=='1':
    if time.localtime()[1]<10:
        mes='0'+str(time.localtime()[1])
    else:
        mes=str(time.localtime()[1])
    date=str(time.localtime()[2])+'.'+mes+'.'+str(time.localtime()[0]-2000)
else:
    date=input('Введите дату (в формате дд.мм.гг)')
proverka=['h','t','t','p',':']
check=0
for i in range(5):
    if proverka[i]!=fak[i]:
        check=1
if check==0:  
#3.Обрезание
#3.1.Скачал картинку
    Im=im.new('RGB',(1000,1000))
    Im.save(name+'.png')
    p = requests.get(fak)
    out = open(name+'.png', "wb")
    out.write(p.content)
    out.close()
else:
    Im=im.open(fak)
    Im.save(name+'.png')
#3.2.Поиск "гнаниц рамки"
image = im.open(name+'.png')
#image.show()
width = image.size[0]
height = image.size[1]
pix=image.load()
Ramka=[2,2,2,2]
i=0
spis2=[width,height,width,height]
g=0
while g in range(4):
    control=Ramka[g]
    for i in range(round(spis2[g]/2)):
            spis1=[[[i+1,round(height/2)],[i,round(height/2)]],
                   [[round(width/2),i+1],[round(width/2),i]],
                   [[width-2-i,round(height/2)],[width-1-i,round(height/2)]],
                   [[round(width/2),height-2-i],[round(width/2),height-1-i]]]
            Ramka[g]+=1
            if control==Ramka[g]-1:
                control=Ramka[g]
                if (abs(pix[spis1[g][0][0],spis1[g][0][1]][0]-176)<10) and (abs(pix[spis1[g][1][0],spis1[g][1][1]][0])<10):
                    Ramka[g]+=1
    Ramka[g]=control
    g+=1
#3.3. Подрезка
area=(Ramka[0],Ramka[1],width-Ramka[2],height-Ramka[3])
image=image.crop(area)
#image.show()
#1.3. Проверяем наличие только нужной "гаммы"
width = image.size[0]
height = image.size[1]
pix=image.load()
n=0
s=[[],[],[],[]]
for i in range(width):
 for j in range(height):
    if pix[i,j][0] not in s[0]:
        s[0].append(pix[i,j][0])
        if abs(176-pix[i,j][0])<20:
            s[2].append(pix[i,j][0])
        if abs(pix[i,j][0])<20:
            s[1].append(pix[i,j][0])
        if abs(255-pix[i,j][0])<20:
            s[3].append(pix[i,j][0])
for i in range(len(s[0])):
    if s[0][i] in s[1] or s[0][i] in s[2] or s[0][i] in s[3]:
        print('ok')
    else:
        print('sosi')
        print(s[0][i])
#4. Заносим в матрицу
pobeda=-1
c=[]
for i in range(width-1):
    #упирамся в серую линию
    if (pix[i+1,1][0] in s[2] and pix[i,j][0] not in s[2]) or (i+1==width-1 and pix[i,j][0] not in s[2]):
        pobeda+=1
        c.append([])
        #Считываем +записываем столбец
        for j in range(height-1):
            if (pix[i,j+1][0] in s[2] and pix[i,j][0] in s[1]) or (j+1==height-1 and pix[i,j][0] in s[1]):
                c[pobeda].append('ch')
            if (pix[i,j+1][0] in s[2] and pix[i,j][0] in s[3]) or (j+1==height-1 and pix[i,j][0] in s[3]):
                c[pobeda].append('b')
gor=len(c)
vert=len(c[0])
#5. Рисуем клетки+пишем цифры
wb = Workbook()
sheet=wb.active
sheet.title='Ответ'
wb.create_sheet('Задание')
sheet1=wb['Задание']
levi=[]
sheet.merge_cells(coor(lvo,veo+1)+':'+coor(lvr+lvo-1,ver+veo))
sheet.merge_cells('C2:'+coor(lvr+lvo+gor-2,2))
sheet1.merge_cells(coor(lvo,veo+1)+':'+coor(lvr+lvo-1,ver+veo))
sheet1.merge_cells('C2:'+coor(lvr+lvo+gor-2,2))
sheet.merge_cells('C3:'+coor(lvr+lvo+gor-2,3))
sheet1.merge_cells('C3:'+coor(lvr+lvo+gor-2,3))
for i in range(len(c[0])):#v и levi[b][0] -контроль первой черной клетки снизу(справа) в столбике (строке)
    #levi[b][1]-отвечает за сдвиг внутри рамки при записи цифр
    levi.append([0,0])
for i in range(len(c)):
    verh=0
    v=0
    for j in range(len(c[0])):
        a=len(c)-i-1
        b=len(c[0])-j-1
        if c[a][b]=='ch':       
            sheet[coor(a+lvo+lvr,b+veo+ver+1)].fill=fill(0)#рисуем
            #А вот работаем с циферками
            if v==0:
                sheet[coor(lvo+lvr+a,veo+ver+verh)]=1
                sheet1[coor(lvo+lvr+a,veo+ver+verh)]=1
                v+=1
            elif c[a][b+1]=='ch':
                sheet[coor(lvo+lvr+a,veo+ver+verh)]=1+sheet[coor(lvo+lvr+a,veo+ver+verh)].value
                sheet1[coor(lvo+lvr+a,veo+ver+verh)]=1+sheet1[coor(lvo+lvr+a,veo+ver+verh)].value
            else:
                verh-=1
                sheet[coor(lvo+lvr+a,veo+ver+verh)]=1
                sheet1[coor(lvo+lvr+a,veo+ver+verh)]=1
            if levi[b][0]==0:
                sheet[coor(lvo+lvr+levi[b][1]-1,b+veo+ver+1)]=1
                sheet1[coor(lvo+lvr+levi[b][1]-1,b+veo+ver+1)]=1
                levi[b][0]+=1
            elif c[a+1][b]=='ch':
                sheet[coor(lvo+lvr+levi[b][1]-1,b+veo+ver+1)]=1+sheet[coor(lvo+lvr+levi[b][1]-1,b+veo+ver+1)].value
                sheet1[coor(lvo+lvr+levi[b][1]-1,b+veo+ver+1)]=1+sheet1[coor(lvo+lvr+levi[b][1]-1,b+veo+ver+1)].value
            else:
                levi[b][1]-=1
                sheet[coor(lvo+lvr+levi[b][1]-1,b+veo+ver+1)]=1
                sheet1[coor(lvo+lvr+levi[b][1]-1,b+veo+ver+1)]=1
#Оформление
for cellObj in sheet[coor(lvo,veo+1)+':'+coor(lvo+lvr+gor-1,ver+veo+vert)]:
    for cell in cellObj:
        q=[0,0,0,0]
        r=[1,1,1,1]
        for i in range(lvo+lvr+gor):
            if cell.column==buk(i):
                b=i
        a=cell.row
        if b in [lvo,lvr+lvo]:
            q[0]=1
            r[0]=0
        if b==lvo+lvr+gor-1:
            q[1]=1
            r[1]=0
        if a in [veo+1,veo+ver+1]:
            q[2]=1
            r[2]=0
        if a==veo+ver+vert:
            q[3]=1
            r[3]=0
        if a>veo+ver and (veo+ver)%5==a%5 and a not in [ver+veo+vert,veo+ver]:
            q[3]=2
        if b>lvo+lvr-1 and (lvo+lvr)%5==b%5 and b not in [lvo+lvr+gor-1,lvo+lvr]:
            q[0]=2
        qq=[]
        rr=[]
        for po in range(4):
            qq.append(q[po])
            rr.append(r[po])
        sheet[cell.coordinate].border=border(q,r)
        sheet[cell.coordinate].alignment=align_center
        sheet1[cell.coordinate].border=border(qq,rr)
        sheet1[cell.coordinate].alignment=align_center
for i in range(gor+lvo+lvr):
    sheet.column_dimensions[sheet[1][i].column].width =2.8
    sheet1.column_dimensions[sheet[1][i].column].width =2.8
sheet['C2'].value='Ответ на кроссворд "'+name+'" от '+date
sheet['C3'].value='Понравилось? - плюсуй сюда:'+'_'*(gor*2+int(round(gor*0.45)))
sheet1['C2'].value=date+' Японский кроссворд "'+name+'"'
sheet1['C3'].value='Ответы нести в комнату...ну где кароче будем'
sheet['C2'].font=font
sheet1['C2'].font=font
sheet['C2'].alignment=align_center
sheet1['C2'].alignment=align_center
#sheet['C3'].alignment=align_center
sheet1['C3'].alignment=align_center
os.remove(name+'.png')
wb.save(name+'.xlsx')                 
                   
        
                
            
            
            
        
        
  

    

    
    

    

        

    
