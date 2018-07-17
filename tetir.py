from PIL import Image as im
import string
import openpyxl
name=input('название экселя в итоге ')
fak=input('путь до файла ')
#wb = openpyxl.load_workbook('E:Кросы\Эндрю.xlsx')
#sheet=wb['Лист1']
#for v in range(100):
#    for t in range(100):
#        sheet.cell(row=v+1, column=t+1).value=''
#wb.save('E:Кросы\Эндрю.xlsx')       
lv=int(input())
ve=int(input())
image = im.open(fak)
width = image.size[0]
height = image.size[1]
print(width)
print(height)
pix=image.load()
n=0
s=[]
sto=[]
nol=[]
dve=[]
c1=[0]
l=string.ascii_uppercase
for i in range(width):
 for j in range(height):
    if pix[i,j][0] not in s:
        s.append(pix[i,j][0])
        if abs(176-pix[i,j][0])<10:
            sto.append(pix[i,j][0])
        if abs(pix[i,j][0])<10:
            nol.append(pix[i,j][0])
        if abs(255-pix[i,j][0])<10:
            dve.append(pix[i,j][0])
for i in range(len(s)):
    if s[i] in sto or s[i] in nol or s[i] in dve:
        print('ok')
    else:
        print('sosi')
        print(s[i])
wb = openpyxl.workbook.Workbook(name)
wb.create_sheet('Лист1')
wb.save(name)
for i in range(width-1):
  c=[]
  c1=[0]
  if (pix[i+1,1][0] in sto and pix[i,j][0] not in sto) or (i+1==width-1 and pix[i,j][0] not in sto):
     for j in range(height-1):
        if (pix[i,j+1][0] in sto and pix[i,j][0] in nol) or (j+1==height-1 and pix[i,j][0] in nol):
            c.append('ch')
        if (pix[i,j+1][0] in sto and pix[i,j][0] in dve) or (j+1==height-1 and pix[i,j][0] in dve):
            c.append('b')
     for f in range(len(c)):
         if c[f]=='ch':
             c1[len(c1)-1]+=1
         else:
             if c1!=[0] and c[f-1]=='ch':
                 c1.append(0)
     if c1[len(c1)-1]==0:
        c1.remove(0)
     print(c1)
     wb = openpyxl.load_workbook(name)
     sheet = wb['Лист1']
     for i in range(len(c1)):
      cif=str(ve-len(c1)+1+i)
      if n+lv>25:
       buk='A'+l[lv+n-26]
      else:
       buk=l[lv+n]
      print(buk+cif)
      sheet[buk+cif]=c1[i]
      wb.save(name)
     n+=1
n=0
for i in range(height-1):
  j=0
  c=[]
  c1=[0]
  if (pix[1,i+1][0] in sto and pix[j,i][0] not in sto) or (i+1==height-1 and pix[j,i][0] not in sto):
     for j in range(width-1):
        if (pix[j+1,i][0] in sto and pix[j,i][0] in nol) or (j+1==width-1 and pix[j,i][0] in nol):
            c.append('ch')
        if (pix[j+1,i][0] in sto and pix[j,i][0] in dve) or (j+1==width-1 and pix[j,i][0] in dve):
            c.append('b')
     for f in range(len(c)):
         if c[f]=='ch':
             c1[len(c1)-1]+=1
         else:
             if c1!=[0] and c[f-1]=='ch':
                 c1.append(0)
     if c1[len(c1)-1]==0:
        c1.remove(0)
     print(c1)
     wb = openpyxl.load_workbook(name)
     sheet = wb['Лист1']
     for i in range(len(c1)):
      cif=str(ve+n+1)
      if lv-len(c1)>25:
       buk='A'+l[lv-len(c1)-26+i]
      else:
       buk=l[lv-len(c1)-26+i]
      sheet[buk+cif]=c1[i]
      wb.save(name)
     n+=1
